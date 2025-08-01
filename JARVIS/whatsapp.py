import pywhatkit
from openpyxl import load_workbook, Workbook
from tkinter import filedialog
import os
from pathlib import Path

WHATSAPP_PATH = Path.home() / "Documents" / "contatos_whatsapp.xlsx"

def carregar_contatos_whatsapp():
    if not os.path.exists(WHATSAPP_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["Nome", "Número WhatsApp"])
        wb.save(WHATSAPP_PATH)

    wb = load_workbook(WHATSAPP_PATH)
    ws = wb.active
    return [(row[0].value, row[1].value) for row in ws.iter_rows(min_row=2, values_only=True) if row[1].value]

def adicionar_contato_whatsapp():
    nome = input("Nome do contato: ")
    numero = input("Número do WhatsApp (com +55...): ")
    
    wb = load_workbook(WHATSAPP_PATH)
    ws = wb.active
    ws.append([nome, numero])
    wb.save(WHATSAPP_PATH)
    print(f"Contato {nome} adicionado com sucesso!")

def gerar_mensagem_com_gemini(prompt):
    # Placeholder - troque pela tua chamada Gemini
    return f"Mensagem gerada a partir do prompt: '{prompt}'"

def enviar_mensagens_whatsapp():
    contatos = carregar_contatos_whatsapp()
    if not contatos:
        print("Nenhum contato encontrado.")
        return

    prompt = input("O que você quer dizer na mensagem? ")
    mensagem = gerar_mensagem_com_gemini(prompt)

    for nome, numero in contatos:
        try:
            print(f"Enviando para {nome} ({numero})...")
            pywhatkit.sendwhatmsg_instantly(numero, mensagem, wait_time=10, tab_close=True)
        except Exception as e:
            print(f"Erro ao enviar para {numero}: {e}")
