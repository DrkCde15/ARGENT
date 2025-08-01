import smtplib
from openpyxl import load_workbook, Workbook
from tkinter import filedialog
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import os
from pathlib import Path

EMAIL_PATH = Path.home() / "Documents" / "contatos_whatsapp.xlsx"

def carregar_contatos_email():
    if not os.path.exists(EMAIL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["Nome", "Email"])
        wb.save(EMAIL_PATH)

    wb = load_workbook(EMAIL_PATH)
    ws = wb.active
    return [(row[0].value, row[1].value) for row in ws.iter_rows(min_row=2, values_only=True) if row[1].value]

def adicionar_contato_email():
    nome = input("Nome do contato: ")
    email = input("Email do contato: ")

    wb = load_workbook(EMAIL_PATH)
    ws = wb.active
    ws.append([nome, email])
    wb.save(EMAIL_PATH)
    print(f"Email de {nome} adicionado com sucesso!")

def gerar_mensagem_com_gemini(prompt):
    # Placeholder - troque pela tua chamada Gemini
    return f"Mensagem gerada a partir do prompt: '{prompt}'"

def selecionar_anexo():
    caminho = filedialog.askopenfilename(title="Selecione o anexo")
    return caminho if caminho else None

def enviar_emails():
    contatos = carregar_contatos_email()
    if not contatos:
        print("Nenhum contato de email encontrado.")
        return

    prompt = input("O que você quer dizer no email? ")
    mensagem = gerar_mensagem_com_gemini(prompt)

    usar_anexo = input("Deseja anexar um arquivo? (s/n): ").lower() == 's'
    anexo_path = selecionar_anexo() if usar_anexo else None

    # Configuração do servidor SMTP (Exemplo: Gmail)
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    remetente = input("Seu email: ")
    senha = input("Senha do email (ou App Password): ")

    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    server.login(remetente, senha)

    for nome, email_destino in contatos:
        msg = MIMEMultipart()
        msg["From"] = remetente
        msg["To"] = email_destino
        msg["Subject"] = "Mensagem automática do JARVIS"
        msg.attach(MIMEText(mensagem, "plain"))

        if anexo_path:
            with open(anexo_path, "rb") as f:
                part = MIMEApplication(f.read(), Name=os.path.basename(anexo_path))
                part["Content-Disposition"] = f'attachment; filename="{os.path.basename(anexo_path)}"'
                msg.attach(part)

        try:
            server.sendmail(remetente, email_destino, msg.as_string())
            print(f"Email enviado para {nome} ({email_destino})")
        except Exception as e:
            print(f"Erro ao enviar para {email_destino}: {e}")

    server.quit()